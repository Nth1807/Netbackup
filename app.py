#!/usr/bin/env python3
"""
NetBackup - Network Device Configuration Backup Tool
Features: SSH backup, GitHub push, Config Diff, Ping check,
          Tags/Groups, Multi-user auth, Auto-cleanup old backups
"""

from flask import Flask, jsonify, request, send_from_directory, Response, session
from flask_cors import CORS
import paramiko
import json, os, io, csv, time, threading, datetime, subprocess, re, difflib, hashlib, socket
from pathlib import Path
from functools import wraps
import logging

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__, static_folder='static')
app.secret_key = os.urandom(32)
CORS(app, supports_credentials=True)

BASE_DIR    = Path(__file__).parent
CONFIG_FILE = BASE_DIR / "config.json"
BACKUP_DIR  = BASE_DIR / "backups"
BACKUP_DIR.mkdir(exist_ok=True)

backup_jobs    = {}
job_id_counter = 0
ping_cache     = {}

DEFAULT_CONFIG = {
    "github": {"token": "", "repo": "", "branch": "main", "commit_message": "Auto backup {date}"},
    "devices": [],
    "tags": [],
    "groups": [],
    "users": [{"username": "admin", "password_hash": hashlib.sha256(b"admin").hexdigest(), "role": "admin"}],
    "cleanup": {"enabled": True, "keep_per_device": 10, "max_age_days": 90},
    "schedule": {"enabled": False, "interval_hours": 24}
}

def load_config():
    if CONFIG_FILE.exists():
        cfg = json.loads(CONFIG_FILE.read_text())
        for k, v in DEFAULT_CONFIG.items():
            cfg.setdefault(k, v)
        return cfg
    return json.loads(json.dumps(DEFAULT_CONFIG))

def save_config(cfg):
    CONFIG_FILE.write_text(json.dumps(cfg, indent=2))

def hash_pw(pw): return hashlib.sha256(pw.encode()).hexdigest()

def require_auth(role=None):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            user = session.get("user")
            if not user: return jsonify({"error": "Unauthorized"}), 401
            if role == "admin" and user.get("role") != "admin": return jsonify({"error": "Forbidden"}), 403
            return f(*args, **kwargs)
        return wrapped
    return decorator

# ── AUTH ──────────────────────────────────────────────────────────────────────
@app.route("/api/auth/login", methods=["POST"])
def login():
    data = request.json or {}
    cfg  = load_config()
    for u in cfg["users"]:
        if u["username"] == data.get("username") and u["password_hash"] == hash_pw(data.get("password","")):
            session["user"] = {"username": u["username"], "role": u["role"]}
            return jsonify({"username": u["username"], "role": u["role"]})
    return jsonify({"error": "Invalid credentials"}), 401

@app.route("/api/auth/logout", methods=["POST"])
def logout():
    session.clear(); return jsonify({"ok": True})

@app.route("/api/auth/me", methods=["GET"])
def me():
    user = session.get("user")
    return jsonify(user) if user else (jsonify({"error": "Not logged in"}), 401)

@app.route("/api/users", methods=["GET"])
@require_auth(role="admin")
def list_users():
    return jsonify([{"username": u["username"], "role": u["role"]} for u in load_config()["users"]])

@app.route("/api/users", methods=["POST"])
@require_auth(role="admin")
def add_user():
    data = request.json or {}
    cfg  = load_config()
    if any(u["username"] == data.get("username") for u in cfg["users"]):
        return jsonify({"error": "Username exists"}), 400
    cfg["users"].append({"username": data["username"], "password_hash": hash_pw(data["password"]), "role": data.get("role","viewer")})
    save_config(cfg); return jsonify({"ok": True})

@app.route("/api/users/<username>", methods=["DELETE"])
@require_auth(role="admin")
def delete_user(username):
    if username == session["user"]["username"]: return jsonify({"error": "Cannot delete yourself"}), 400
    cfg = load_config()
    cfg["users"] = [u for u in cfg["users"] if u["username"] != username]
    save_config(cfg); return jsonify({"ok": True})

@app.route("/api/users/<username>/password", methods=["PUT"])
@require_auth(role="admin")
def change_password(username):
    cfg = load_config()
    for u in cfg["users"]:
        if u["username"] == username:
            u["password_hash"] = hash_pw(request.json["password"])
            save_config(cfg); return jsonify({"ok": True})
    return jsonify({"error": "Not found"}), 404

# ── TAGS & GROUPS ─────────────────────────────────────────────────────────────
@app.route("/api/tags", methods=["GET"])
@require_auth()
def list_tags():
    cfg = load_config()
    tags = set(cfg.get("tags", []))
    for d in cfg["devices"]:
        tags.update(d.get("tags", []))
    return jsonify(sorted(tags))

@app.route("/api/groups", methods=["GET"])
@require_auth()
def list_groups():
    cfg = load_config()
    groups = cfg.get("groups", [])
    for g in groups:
        g["device_count"] = sum(1 for d in cfg["devices"] if g["tag"] in d.get("tags", []))
    return jsonify(groups)

@app.route("/api/groups", methods=["POST"])
@require_auth(role="admin")
def add_group():
    data = request.json or {}
    cfg  = load_config()
    grp  = {"id": f"grp_{int(time.time()*1000)}", "name": data["name"], "tag": data["tag"]}
    cfg.setdefault("groups", []).append(grp)
    save_config(cfg); return jsonify(grp)

@app.route("/api/groups/<grp_id>", methods=["DELETE"])
@require_auth(role="admin")
def delete_group(grp_id):
    cfg = load_config()
    cfg["groups"] = [g for g in cfg.get("groups", []) if g["id"] != grp_id]
    save_config(cfg); return jsonify({"ok": True})

# ── PING ──────────────────────────────────────────────────────────────────────
def tcp_ping(host, port=22, timeout=2.0):
    t0 = time.time()
    try:
        with socket.create_connection((host, int(port)), timeout=timeout):
            return {"alive": True, "latency_ms": round((time.time()-t0)*1000, 1)}
    except Exception:
        return {"alive": False, "latency_ms": None}

def ping_all_devices():
    while True:
        cfg = load_config()
        for dev in cfg["devices"]:
            r = tcp_ping(dev["host"], dev.get("port", 22))
            ping_cache[dev["host"]] = {**r, "checked_at": datetime.datetime.now().isoformat()}
        time.sleep(30)

@app.route("/api/ping", methods=["GET"])
@require_auth()
def get_ping_status(): return jsonify(ping_cache)

@app.route("/api/ping/<dev_id>", methods=["POST"])
@require_auth()
def ping_one(dev_id):
    dev = next((d for d in load_config()["devices"] if d["id"] == dev_id), None)
    if not dev: return jsonify({"error": "Not found"}), 404
    r = tcp_ping(dev["host"], dev.get("port", 22))
    ping_cache[dev["host"]] = {**r, "checked_at": datetime.datetime.now().isoformat()}
    return jsonify(ping_cache[dev["host"]])

# ── SSH BACKUP ────────────────────────────────────────────────────────────────
DEVICE_COMMANDS = {
    "cisco_ios":  ["show running-config", "show version", "show ip interface brief"],
    "cisco_nxos": ["show running-config", "show version", "show interface brief"],
    "juniper":    ["show configuration | display text", "show version", "show interfaces terse"],
    "fortigate":  ["show full-configuration", "get system status"],
    "mikrotik":   ["export verbose", "/system resource print"],
    "generic":    ["show running-config"],
}

def ssh_backup(device):
    host, port = device["host"], int(device.get("port", 22))
    result = {"host": host, "status": "error", "output": {}, "error": "", "timestamp": datetime.datetime.now().isoformat()}
    try:
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        kw = dict(hostname=host, port=port, username=device["username"], timeout=30, banner_timeout=30)
        if device.get("key_path"): kw["key_filename"] = device["key_path"]
        else: kw.update(password=device.get("password",""), look_for_keys=False, allow_agent=False)
        client.connect(**kw)
        outputs = {}
        for cmd in DEVICE_COMMANDS.get(device.get("type","generic"), ["show running-config"]):
            try:
                _, stdout, stderr = client.exec_command(cmd, timeout=30)
                out = stdout.read().decode("utf-8", errors="replace")
                outputs[cmd] = out or f"[STDERR] {stderr.read().decode()}"
            except Exception as e: outputs[cmd] = f"[ERROR] {e}"
        client.close()
        result.update({"status": "success", "output": outputs})
    except paramiko.AuthenticationException: result["error"] = "Authentication failed"
    except paramiko.NoValidConnectionsError: result["error"] = f"Cannot connect to {host}:{port}"
    except Exception as e: result["error"] = str(e)
    return result

def save_backup_file(device_name, result):
    ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = re.sub(r'[^a-zA-Z0-9_-]', '_', device_name)
    path = BACKUP_DIR / f"{safe}_{ts}.txt"
    lines = [f"# NetBackup - {device_name}", f"# Timestamp : {result['timestamp']}",
             f"# Status    : {result['status']}", f"# Host      : {result['host']}", "="*70, ""]
    for cmd, out in result.get("output", {}).items():
        lines += [f"\n### Command: {cmd}\n", out, ""]
    if result.get("error"): lines += [f"\n### ERROR\n{result['error']}"]
    path.write_text("\n".join(lines), encoding="utf-8")
    return path

# ── DIFF ──────────────────────────────────────────────────────────────────────
@app.route("/api/diff", methods=["POST"])
@require_auth()
def config_diff():
    data = request.json or {}
    pa, pb = BACKUP_DIR / data.get("file_a",""), BACKUP_DIR / data.get("file_b","")
    if not pa.exists() or not pb.exists(): return jsonify({"error": "File not found"}), 404
    la = pa.read_text(encoding="utf-8", errors="replace").splitlines(keepends=True)
    lb = pb.read_text(encoding="utf-8", errors="replace").splitlines(keepends=True)
    diff = list(difflib.unified_diff(la, lb, fromfile=data["file_a"], tofile=data["file_b"], lineterm=""))
    return jsonify({"diff": diff,
                    "added":   sum(1 for l in diff if l.startswith("+") and not l.startswith("+++")),
                    "removed": sum(1 for l in diff if l.startswith("-") and not l.startswith("---")),
                    "changed": bool(diff)})


# ── CLEANUP ───────────────────────────────────────────────────────────────────
def cleanup_old_backups(cfg=None):
    if cfg is None: cfg = load_config()
    c        = cfg.get("cleanup", DEFAULT_CONFIG["cleanup"])
    keep     = int(c.get("keep_per_device", 10))
    max_days = int(c.get("max_age_days", 90))
    cutoff   = datetime.datetime.now() - datetime.timedelta(days=max_days)
    deleted  = []
    all_files = sorted(BACKUP_DIR.glob("*.txt"), key=lambda f: f.stat().st_mtime, reverse=True)
    by_device = {}
    for f in all_files:
        m = re.match(r'^(.+)_\d{8}_\d{6}\.txt$', f.name)
        by_device.setdefault(m.group(1) if m else f.stem, []).append(f)
    for _, files in by_device.items():
        for i, f in enumerate(files):
            if datetime.datetime.fromtimestamp(f.stat().st_mtime) < cutoff or i >= keep:
                f.unlink(); deleted.append(f.name)
    return {"deleted": deleted, "count": len(deleted)}

@app.route("/api/cleanup", methods=["POST"])
@require_auth(role="admin")
def run_cleanup(): return jsonify(cleanup_old_backups())

@app.route("/api/cleanup/settings", methods=["GET"])
@require_auth()
def get_cleanup_settings(): return jsonify(load_config().get("cleanup", DEFAULT_CONFIG["cleanup"]))

@app.route("/api/cleanup/settings", methods=["PUT"])
@require_auth(role="admin")
def update_cleanup_settings():
    cfg = load_config()
    cfg["cleanup"].update(request.json)
    save_config(cfg); return jsonify(cfg["cleanup"])

# ── GITHUB ────────────────────────────────────────────────────────────────────
def git_push(files, cfg):
    token, repo, branch = cfg["github"]["token"], cfg["github"]["repo"], cfg["github"]["branch"]
    msg = cfg["github"]["commit_message"].replace("{date}", datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
    if not token or not repo: return {"success": False, "error": "GitHub not configured"}
    work_dir = BASE_DIR / "git_work"
    try:
        repo_url = f"https://{token}@github.com/{repo}.git"
        if work_dir.exists():
            subprocess.run(["git","-C",str(work_dir),"pull","--rebase"], check=True, capture_output=True)
        else:
            subprocess.run(["git","clone","--depth=1","-b",branch,repo_url,str(work_dir)], check=True, capture_output=True)
        bd = work_dir/"backups"; bd.mkdir(exist_ok=True)
        pushed = []
        for f in files:
            dest = bd/f.name; dest.write_bytes(f.read_bytes())
            pushed.append(str(dest.relative_to(work_dir)))
        subprocess.run(["git","-C",str(work_dir),"add","."], check=True, capture_output=True)
        r = subprocess.run(["git","-C",str(work_dir),"diff","--cached","--quiet"], capture_output=True)
        if r.returncode != 0:
            subprocess.run(["git","-C",str(work_dir),"commit","-m",msg], check=True, capture_output=True)
            subprocess.run(["git","-C",str(work_dir),"push","origin",branch], check=True, capture_output=True)
            return {"success": True, "files": pushed, "message": msg}
        return {"success": True, "files": [], "message": "No changes"}
    except subprocess.CalledProcessError as e: return {"success": False, "error": e.stderr.decode() if e.stderr else str(e)}
    except Exception as e: return {"success": False, "error": str(e)}

# ── BACKUP JOB ────────────────────────────────────────────────────────────────
def run_backup_job(job_id, device_ids):
    global backup_jobs
    cfg = load_config()
    devices = {d["id"]: d for d in cfg["devices"]}
    backup_jobs[job_id]["status"] = "running"
    backed_files, results = [], []
    for dev_id in device_ids:
        dev = devices.get(dev_id)
        if not dev: continue
        backup_jobs[job_id]["current"] = dev["name"]
        result = ssh_backup(dev)
        path   = save_backup_file(dev["name"], result)
        backed_files.append(path)
        results.append({"device": dev["name"], "status": result["status"], "error": result.get("error",""), "file": path.name})
        cfg = load_config()
        for d in cfg["devices"]:
            if d["id"] == dev_id:
                d["last_backup"] = result["timestamp"]; d["last_status"] = result["status"]; break
        save_config(cfg)
    git_result = {"success": False, "error": "Skipped"}
    if backed_files and cfg["github"]["token"]: git_result = git_push(backed_files, cfg)
    if cfg.get("cleanup", {}).get("enabled", True): cleanup_old_backups(cfg)
    backup_jobs[job_id].update({"status": "done", "results": results, "git": git_result,
                                 "finished_at": datetime.datetime.now().isoformat()})

# ── DEVICES API ───────────────────────────────────────────────────────────────
@app.route("/api/config", methods=["GET"])
@require_auth()
def get_config():
    cfg = load_config(); safe = json.loads(json.dumps(cfg))
    if safe["github"]["token"]: safe["github"]["token"] = "***" + safe["github"]["token"][-4:]
    safe.pop("users", None); return jsonify(safe)

@app.route("/api/config", methods=["POST"])
@require_auth(role="admin")
def update_config():
    data = request.json; cfg = load_config()
    if "github" in data:
        for k, v in data["github"].items():
            if k == "token" and str(v).startswith("***"): continue
            cfg["github"][k] = v
    if "schedule" in data: cfg["schedule"].update(data["schedule"])
    save_config(cfg); return jsonify({"ok": True})

@app.route("/api/devices", methods=["GET"])
@require_auth()
def list_devices():
    cfg = load_config(); devs = cfg["devices"]
    tag = request.args.get("tag")
    if tag: devs = [d for d in devs if tag in d.get("tags", [])]
    for d in devs: d["ping"] = ping_cache.get(d["host"])
    return jsonify(devs)

@app.route("/api/devices", methods=["POST"])
@require_auth(role="admin")
def add_device():
    cfg = load_config(); dev = request.json
    dev.update({"id": f"dev_{int(time.time()*1000)}", "last_backup": None, "last_status": None})
    dev.setdefault("tags", [])
    cfg["devices"].append(dev); save_config(cfg); return jsonify(dev)

@app.route("/api/devices/<dev_id>", methods=["PUT"])
@require_auth(role="admin")
def update_device(dev_id):
    cfg = load_config()
    for i, d in enumerate(cfg["devices"]):
        if d["id"] == dev_id:
            cfg["devices"][i].update(request.json); save_config(cfg); return jsonify(cfg["devices"][i])
    return jsonify({"error": "Not found"}), 404

@app.route("/api/devices/<dev_id>", methods=["DELETE"])
@require_auth(role="admin")
def delete_device(dev_id):
    cfg = load_config()
    cfg["devices"] = [d for d in cfg["devices"] if d["id"] != dev_id]
    save_config(cfg); return jsonify({"ok": True})

VENDOR_MAP = {"cisco":"cisco_ios","cisco ios":"cisco_ios","cisco_ios":"cisco_ios","cisco nxos":"cisco_nxos",
              "cisco_nxos":"cisco_nxos","nxos":"cisco_nxos","nexus":"cisco_nxos","juniper":"juniper",
              "junos":"juniper","fortigate":"fortigate","fortinet":"fortigate","mikrotik":"mikrotik","generic":"generic"}

def normalize_vendor(v): return VENDOR_MAP.get(v.strip().lower(), "generic")

def parse_rows(rows):
    cfg = load_config(); existing = {d["host"] for d in cfg["devices"]}
    imported, errors = [], []
    for i, row in enumerate(rows, start=2):
        name  = (row.get("name") or row.get("hostname") or "").strip()
        host  = (row.get("management_ip") or row.get("ip") or row.get("host") or "").strip()
        user  = (row.get("user") or row.get("username") or "").strip()
        pwd   = (row.get("pass") or row.get("password") or "").strip()
        port  = str(row.get("port") or "22").strip() or "22"
        tags  = [t.strip() for t in str(row.get("tags") or row.get("site") or "").split(",") if t.strip()]
        dtype = normalize_vendor(str(row.get("vendor") or row.get("type") or "generic"))
        if not name or not host: errors.append({"row": i, "reason": "Missing name or IP"}); continue
        if not user: errors.append({"row": i, "reason": f"Missing username for {name}"}); continue
        if host in existing: errors.append({"row": i, "reason": f"{name} ({host}) already exists"}); continue
        try: port_int = int(port)
        except: port_int = 22
        dev = {"id": f"dev_{int(time.time()*1000)}_{i}", "name": name, "host": host, "port": port_int,
               "username": user, "password": pwd, "key_path": "", "type": dtype, "tags": tags,
               "last_backup": None, "last_status": None}
        cfg["devices"].append(dev); existing.add(host)
        imported.append({"name": name, "host": host, "type": dtype, "tags": tags})
    save_config(cfg); return imported, errors

@app.route("/api/devices/import", methods=["POST"])
@require_auth(role="admin")
def import_devices():
    if "file" not in request.files: return jsonify({"error": "No file"}), 400
    f = request.files["file"]; fname = f.filename.lower(); rows = []
    try:
        if fname.endswith(".csv"):
            rows = [dict(r) for r in csv.DictReader(io.StringIO(f.read().decode("utf-8-sig")))]
        elif fname.endswith((".xlsx",".xls")):
            if not HAS_OPENPYXL: return jsonify({"error": "pip install openpyxl"}), 500
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True); ws = wb.active
            hdrs = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1,max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row): continue
                rows.append({hdrs[i]: (str(v).strip() if v is not None else "") for i,v in enumerate(row)})
        else: return jsonify({"error": "Only .csv or .xlsx"}), 400
        imported, errors = parse_rows(rows)
        return jsonify({"imported": imported, "errors": errors, "total_rows": len(rows)})
    except Exception as e: logger.exception("Import error"); return jsonify({"error": str(e)}), 500

@app.route("/api/devices/import/template")
def download_template():
    c = "name,management_ip,vendor,user,pass,port,tags\nCore-SW-01,192.168.1.1,cisco_ios,admin,P@ss,22,HCM-DC\nFW-01,10.10.0.1,fortigate,admin,,22,HN-DC\n"
    return Response(c, mimetype="text/csv", headers={"Content-Disposition":"attachment; filename=devices_template.csv"})

@app.route("/api/backup", methods=["POST"])
@require_auth()
def start_backup():
    global job_id_counter
    data = request.json or {}; cfg = load_config()
    device_ids = data.get("device_ids") or [d["id"] for d in cfg["devices"]]
    job_id_counter += 1; job_id = f"job_{job_id_counter}"
    backup_jobs[job_id] = {"id": job_id, "status": "queued", "current": "", "results": [], "git": {},
                            "started_at": datetime.datetime.now().isoformat(), "device_ids": device_ids,
                            "started_by": session.get("user",{}).get("username","?")}
    threading.Thread(target=run_backup_job, args=(job_id, device_ids), daemon=True).start()
    return jsonify({"job_id": job_id})

@app.route("/api/jobs", methods=["GET"])
@require_auth()
def list_jobs(): return jsonify(list(reversed(list(backup_jobs.values())))[:20])

@app.route("/api/jobs/<job_id>", methods=["GET"])
@require_auth()
def get_job(job_id):
    j = backup_jobs.get(job_id)
    return jsonify(j) if j else (jsonify({"error":"Not found"}),404)

@app.route("/api/backups", methods=["GET"])
@require_auth()
def list_backups():
    files = sorted(BACKUP_DIR.glob("*.txt"), key=lambda f: f.stat().st_mtime, reverse=True)
    return jsonify([{"name":f.name,"size":f.stat().st_size,
                     "modified":datetime.datetime.fromtimestamp(f.stat().st_mtime).isoformat()} for f in files[:100]])

@app.route("/api/backups/<filename>", methods=["GET"])
@require_auth()
def get_backup_file(filename): return send_from_directory(str(BACKUP_DIR), filename, as_attachment=True)

@app.route("/api/backups/<filename>", methods=["DELETE"])
@require_auth(role="admin")
def delete_backup_file(filename):
    p = BACKUP_DIR/filename
    if p.exists(): p.unlink(); return jsonify({"ok":True})
    return jsonify({"error":"Not found"}),404

@app.route("/api/backups/by-device/<device_name>", methods=["GET"])
@require_auth()
def backups_by_device(device_name):
    safe  = re.sub(r'[^a-zA-Z0-9_-]','_',device_name)
    files = sorted(BACKUP_DIR.glob(f"{safe}_*.txt"), key=lambda f: f.stat().st_mtime, reverse=True)
    return jsonify([{"name":f.name,"size":f.stat().st_size,
                     "modified":datetime.datetime.fromtimestamp(f.stat().st_mtime).isoformat()} for f in files])

@app.route("/api/stats", methods=["GET"])
@require_auth()
def get_stats():
    cfg = load_config(); devs = cfg["devices"]; files = list(BACKUP_DIR.glob("*.txt"))
    return jsonify({"total_devices":len(devs),
                    "ok":sum(1 for d in devs if d.get("last_status")=="success"),
                    "failed":sum(1 for d in devs if d.get("last_status")=="error"),
                    "never_backed":sum(1 for d in devs if not d.get("last_status")),
                    "alive_devices":sum(1 for d in devs if ping_cache.get(d["host"],{}).get("alive")),
                    "total_backups":len(files),
                    "total_size_kb":round(sum(f.stat().st_size for f in files)/1024,1),
                    "last_job":list(backup_jobs.values())[-1] if backup_jobs else None})

@app.route("/", defaults={"path":""})
@app.route("/<path:path>")
def serve(path): return send_from_directory(str(BASE_DIR/"static"), "index.html")

if __name__ == "__main__":
    threading.Thread(target=ping_all_devices, daemon=True).start()
    logger.info("NetBackup on http://localhost:5000  |  login: admin / admin")
    app.run(host="0.0.0.0", port=5000, debug=False)
